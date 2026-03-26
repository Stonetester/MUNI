import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\keato\Downloads\KK Finances_ Rework.xlsx', data_only=True)

print("=== SHEET NAMES ===")
for i, name in enumerate(wb.sheetnames):
    print(f"  {i}: {name!r}")

# ── Keaton PayStub ──
ws = wb['Keaton PayStub']
print("\n=== KEATON PAYSTUB rows 1-25, cols A-P ===")
for r in ws.iter_rows(min_row=1, max_row=25, min_col=1, max_col=16, values_only=True):
    print(r)

print("\n=== KEATON PAYSTUB rows 1-25, cols Q-AF (tax breakdown) ===")
for r in ws.iter_rows(min_row=1, max_row=25, min_col=17, max_col=32, values_only=True):
    print(r)

# ── Katherine PayStub ──
ws2 = wb['Katherine PayStub']
print("\n=== KATHERINE PAYSTUB rows 1-25, cols A-P ===")
for r in ws2.iter_rows(min_row=1, max_row=25, min_col=1, max_col=16, values_only=True):
    print(r)

print("\n=== KATHERINE PAYSTUB rows 1-25, cols Q-AF ===")
for r in ws2.iter_rows(min_row=1, max_row=25, min_col=17, max_col=32, values_only=True):
    print(r)

# ── KD Ongoing Tracker (first 6 cols to see row labels + month headers) ──
ws3 = wb['KD Ongoing Tracker']
print("\n=== KD ONGOING TRACKER rows 1-60, cols A-G ===")
for r in ws3.iter_rows(min_row=1, max_row=60, min_col=1, max_col=7, values_only=True):
    print(r)

# ── KAK Ongoing Tracker ──
ws4 = wb['KAK Ongoing Tracker']
print("\n=== KAK ONGOING TRACKER rows 1-60, cols A-G ===")
for r in ws4.iter_rows(min_row=1, max_row=60, min_col=1, max_col=7, values_only=True):
    print(r)
