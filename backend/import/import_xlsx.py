"""
FinanceTrack -- One-Time Historical Data Import
Reads KK Finances_ Rework.xlsx and posts historical data to the FinanceTrack API.

Account IDs and category IDs are resolved automatically by querying the API --
no manual ID lookup or account_map.json editing required.

Usage:
    python import_xlsx.py --dry-run          # Preview without writing anything
    python import_xlsx.py                    # Full import (both users)
    python import_xlsx.py --user keaton      # Keaton only
    python import_xlsx.py --user katherine   # Katherine only
    python import_xlsx.py --skip-paystubs    # Skip paystubs
    python import_xlsx.py --skip-snapshots   # Skip balance snapshots
    python import_xlsx.py --skip-transactions  # Skip income transactions
    python import_xlsx.py --skip-loans       # Skip student loan payment transactions
    python import_xlsx.py --skip-bonuses     # Skip bonus income transactions

What gets imported:
    - Keaton paystubs (7 records: Feb-Aug 2025)
    - Katherine paystub (1 record: Mar 2025)
    - Keaton balance snapshots (Jan + Mar 2025)
    - Monthly income totals Sep 2024 - Jan 2025 (KD/KAK Ongoing Tracker row 20)
    - Student loan payments Sep 2024 - Jan 2025 (KD row 132 / KAK row 108)
      NOTE: Student loan account must exist in the app before running.
            The import posts payments as expense transactions. Set your starting
            balance as a balance snapshot manually in the app.
    - Bonus income, all months with a value (KD row 30 / KAK rows 26-27)
      Bonuses are manual entries in the Rework sheet not captured in Google Sheets.

Prerequisites:
    1. Backend must be running on localhost:8000
    2. Accounts must exist in the app (created in UI first):
       - Student Loans account (type: student_loan) for each user who had loans
    3. pip install openpyxl requests
"""

import argparse
import json
import re
import sys
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl
import requests

# -----------------------------------------------------------------------------
# Config -- only path/URL settings, no IDs needed
# -----------------------------------------------------------------------------

SCRIPT_DIR = Path(__file__).parent
CONFIG_FILE = SCRIPT_DIR / "config.json"

DEFAULT_CONFIG = {
    "xlsx_path": r"C:\Users\keato\Downloads\KK Finances_ Rework.xlsx",
    "api_base":  "http://localhost:8000/api/v1",
    "keaton_username":    "keaton",
    "keaton_password":    "finance123",
    "katherine_username": "katherine",
    "katherine_password": "finance123",
}

# Tax breakdown reference extracted from Keaton PayStub right-side table:
# gross -> (state_md, county_md_cal, federal, medicare, social_security)
KEATON_TAX_REF = {
    3635.41: (141.22, 95.14, 330.33, 52.69, 225.29),
    4877.53: (203.17, 136.96, 616.02, 70.70, 302.30),
}


# -----------------------------------------------------------------------------
# Date Parsing
# -----------------------------------------------------------------------------

def parse_pay_period(cell_value, default_year: int = 2025):
    """
    Parse pay period strings into (period_start, period_end, pay_date).

    Handles:
      '3/16-3/31'   -> (2025-03-16, 2025-03-31, 2025-03-31)
      '2/16 -2/28'  -> (2025-02-16, 2025-02-28, 2025-02-28)
      datetime obj  -> (that date, that date, that date)

    Returns None if not parseable or empty.
    """
    if cell_value is None:
        return None
    if isinstance(cell_value, (datetime, date)):
        d = cell_value.date() if isinstance(cell_value, datetime) else cell_value
        return d, d, d
    if not isinstance(cell_value, str):
        return None
    cell_value = cell_value.strip()
    if not cell_value:
        return None
    m = re.match(r'^(\d{1,2})/(\d{1,2})\s*[-\u2013]\s*(\d{1,2})/(\d{1,2})$', cell_value)
    if not m:
        return None
    s_mo, s_day, e_mo, e_day = (int(m.group(i)) for i in range(1, 5))
    try:
        start = date(default_year, s_mo, s_day)
        end   = date(default_year, e_mo, e_day)
        return start, end, end
    except ValueError:
        return None


def last_day_of_month(year: int, month: int) -> date:
    """Return the last calendar day of a given month."""
    if month == 12:
        return date(year, 12, 31)
    return date(year, month + 1, 1) - timedelta(days=1)


# -----------------------------------------------------------------------------
# API -- auto-discovery
# -----------------------------------------------------------------------------

def login(api_base: str, username: str, password: str) -> str:
    """Login and return the JWT access token."""
    resp = requests.post(
        f"{api_base}/auth/login",
        data={"username": username, "password": password},
        headers={"Content-Type": "application/x-www-form-urlencoded"},
        timeout=10,
    )
    resp.raise_for_status()
    return resp.json()["access_token"]


def discover_accounts(api_base: str, token: str) -> dict:
    """
    Query the API and return {account_type: account_id} for the logged-in user.

    Example return value:
        {
            "checking": 1,
            "savings":  2,
            "hysa":     3,
            "401k":     4,
            "ira":      5,
        }

    Account types not found in the app return no entry (caller checks with .get()).
    """
    resp = requests.get(
        f"{api_base}/accounts",
        headers={"Authorization": f"Bearer {token}"},
        timeout=10,
    )
    resp.raise_for_status()
    accounts = resp.json()
    result = {}
    for acct in accounts:
        atype = acct.get("account_type", "").lower()
        if atype and atype not in result:
            result[atype] = acct["id"]
    return result


def discover_income_category(api_base: str, token: str) -> int | None:
    """
    Query categories and return the ID of the first category with kind='income'.
    Returns None if not found.
    """
    resp = requests.get(
        f"{api_base}/categories",
        headers={"Authorization": f"Bearer {token}"},
        timeout=10,
    )
    resp.raise_for_status()
    for cat in resp.json():
        if cat.get("kind") == "income":
            return cat["id"]
    return None


def discover_category_by_name(api_base: str, token: str, name: str) -> int | None:
    """Return the category ID matching the given name (case-insensitive). None if not found."""
    resp = requests.get(
        f"{api_base}/categories",
        headers={"Authorization": f"Bearer {token}"},
        timeout=10,
    )
    resp.raise_for_status()
    name_lower = name.lower()
    for cat in resp.json():
        if cat.get("name", "").lower() == name_lower:
            return cat["id"]
    return None


def get_existing_transaction_keys(api_base: str, token: str) -> set:
    """
    Return set of (date_str, description) tuples for all existing xlsx-imported transactions.
    Used to prevent duplicate imports on re-runs.
    """
    resp = requests.get(
        f"{api_base}/transactions",
        params={"limit": 2000, "offset": 0},
        headers={"Authorization": f"Bearer {token}"},
        timeout=10,
    )
    if resp.status_code == 404:
        return set()
    resp.raise_for_status()
    data = resp.json()
    items = data.get("items", data) if isinstance(data, dict) else data
    return {
        (item["date"], item["description"])
        for item in items
        if "(xlsx import)" in item.get("description", "")
    }


def get_existing_paystub_dates(api_base: str, token: str) -> set:
    """Return set of existing pay_date strings to avoid duplicates."""
    resp = requests.get(
        f"{api_base}/paystubs",
        headers={"Authorization": f"Bearer {token}"},
        timeout=10,
    )
    if resp.status_code == 404:
        return set()
    resp.raise_for_status()
    items = resp.json()
    if isinstance(items, dict):
        items = items.get("items", items.get("paystubs", []))
    return {item["pay_date"] for item in items}


def get_existing_snapshot_dates(api_base: str, token: str, account_id: int) -> set:
    """Return set of existing snapshot date strings for an account."""
    resp = requests.get(
        f"{api_base}/balance-snapshots",
        params={"account_id": account_id},
        headers={"Authorization": f"Bearer {token}"},
        timeout=10,
    )
    if resp.status_code == 404:
        return set()
    resp.raise_for_status()
    items = resp.json()
    if isinstance(items, dict):
        items = items.get("items", [])
    return {item["date"] for item in items}


# -----------------------------------------------------------------------------
# Excel Extraction
# -----------------------------------------------------------------------------

def extract_keaton_paystubs(ws) -> list:
    """
    Extract paystub records from the 'Keaton PayStub' sheet.

    Sheet layout (1-indexed rows):
      Row 5  = most recent paystub
      Row 7  = second header row (skip)
      Rows 8-14 = historical paystubs

    Column mapping (0-indexed from A):
      B(1)=date string, C(2)=income/gross_pay, D(3)=taxes_total,
      E(4)=deductions_total, F(5)=401k_employee, G(6)=benefits, H(7)=net_pay

    Naming quirk in spreadsheet: 'income' column = GROSS pay; 'gross income' = NET pay.
    Verified: income - taxes - deductions = gross_income (e.g. 3635.41 - 844.67 - 562.31 = 2228.43)
    """
    records = []
    data_rows = [5] + list(range(8, 15))
    seen_dates = set()

    for row_num in data_rows:
        row = [ws.cell(row_num, col).value for col in range(1, 17)]
        date_cell = row[1]
        parsed = parse_pay_period(date_cell)
        if parsed is None:
            continue

        period_start, period_end, pay_date = parsed
        if pay_date in seen_dates:
            continue
        seen_dates.add(pay_date)

        gross_pay      = _num(row[2])
        tax_total      = _num(row[3])
        deduction_tot  = _num(row[4])
        deduction_401k = _num(row[5])
        net_pay        = _num(row[7])

        # Individual tax breakdown from reference table; scale for unknown pay rates
        tax_ref = KEATON_TAX_REF.get(gross_pay)
        if tax_ref:
            md, md_cal, fed, med, ss = tax_ref
        elif gross_pay:
            nearest = min(KEATON_TAX_REF, key=lambda k: abs(k - gross_pay))
            ratio = gross_pay / nearest
            md, md_cal, fed, med, ss = (round(v * ratio, 2) for v in KEATON_TAX_REF[nearest])
        else:
            md = md_cal = fed = med = ss = 0.0

        records.append({
            "pay_date":            pay_date.isoformat(),
            "period_start":        period_start.isoformat(),
            "period_end":          period_end.isoformat(),
            "employer":            "ACE (xlsx import)",
            "gross_pay":           gross_pay,
            "regular_pay":         gross_pay,
            "tax_total":           tax_total,
            "tax_state":           md,
            "tax_county":          md_cal,
            "tax_federal":         fed,
            "tax_medicare":        med,
            "tax_social_security": ss,
            "deduction_total":     deduction_tot,
            "deduction_401k":      deduction_401k,
            "net_pay":             net_pay,
            "parse_method":        "xlsx_import",
            "notes":               f"Imported from KK Finances_ Rework.xlsx ({date_cell})",
        })

    return records


def extract_keaton_snapshots(ws, account_ids: dict) -> list:
    """
    Extract balance snapshots from 'Keaton PayStub' sheet.

    account_ids is the auto-discovered {account_type: id} dict.
    Columns K-O (1-indexed 11-15) hold balances:
      K=checking, L=savings, M=hysa, N=401k, O=ira

    Only creates snapshots for account types that exist in the app.
    """
    col_account = [
        (10, account_ids.get("checking"), "Checking"),
        (11, account_ids.get("savings"),  "Savings"),
        (12, account_ids.get("hysa"),     "HYSA"),
        (13, account_ids.get("401k"),     "401k"),
        (14, account_ids.get("ira"),      "IRA"),
    ]
    snapshots = []

    # Row 5 = most recent paystub row -- use period_end as snapshot date
    row5_date_cell = ws.cell(5, 2).value
    parsed5 = parse_pay_period(row5_date_cell)
    if parsed5:
        snap_date = parsed5[2]
        row5 = [ws.cell(5, col).value for col in range(1, 16)]
        for col_idx, acct_id, label in col_account:
            if acct_id is None:
                continue
            bal = _num(row5[col_idx])
            if bal is not None:
                snapshots.append({
                    "account_id": acct_id,
                    "date":       snap_date.isoformat(),
                    "balance":    bal,
                    "notes":      f"From paystub {row5_date_cell} (xlsx import)",
                })

    # Rows 8-14 -- only snapshot when Month label (col J = col 10 in 1-indexed) is set
    for row_num in range(8, 15):
        month_label = ws.cell(row_num, 10).value
        if not month_label:
            continue

        date_cell = ws.cell(row_num, 2).value
        parsed = parse_pay_period(date_cell)
        if parsed is None:
            continue

        # Month label (e.g. 'Jan') = the month these are end-of-month balances for
        # Row labeled 'Jan' has period '2/16-2/28' -> snapshot date = Jan 31
        period_start = parsed[0]
        snap_month = period_start.month - 1 if period_start.month > 1 else 12
        snap_year  = period_start.year if period_start.month > 1 else period_start.year - 1
        snap_date  = last_day_of_month(snap_year, snap_month)

        row = [ws.cell(row_num, col).value for col in range(1, 16)]
        for col_idx, acct_id, label in col_account:
            if acct_id is None:
                continue
            bal = _num(row[col_idx])
            if bal is not None:
                snapshots.append({
                    "account_id": acct_id,
                    "date":       snap_date.isoformat(),
                    "balance":    bal,
                    "notes":      f"Month: {month_label} (xlsx import)",
                })

    return snapshots


def extract_katherine_paystubs(ws) -> list:
    """
    Extract paystub records from 'Katherine PayStub' sheet.

    Column mapping (0-indexed from A=0):
      A(0)=Date, B(1)=income(gross), C(2)=taxes, D(3)=deductions,
      E(4)=401k, F(5)=benefits, G(6)=net_pay
    """
    records = []
    seen_dates = set()

    for row_num in [5, 8]:
        row = [ws.cell(row_num, col).value for col in range(1, 9)]
        date_cell = row[0]
        parsed = parse_pay_period(date_cell)
        if parsed is None:
            continue

        period_start, period_end, pay_date = parsed
        if pay_date in seen_dates:
            continue
        seen_dates.add(pay_date)

        gross_pay      = _num(row[1])
        tax_total      = _num(row[2])
        deduction_tot  = _num(row[3])
        deduction_401k = _num(row[4])
        net_pay        = _num(row[6])

        records.append({
            "pay_date":        pay_date.isoformat(),
            "period_start":    period_start.isoformat(),
            "period_end":      period_end.isoformat(),
            "employer":        "G&P (xlsx import)",
            "gross_pay":       gross_pay,
            "regular_pay":     gross_pay,
            "tax_total":       tax_total,
            "deduction_total": deduction_tot,
            "deduction_401k":  deduction_401k,
            "net_pay":         net_pay,
            "parse_method":    "xlsx_import",
            "notes":           f"Imported from KK Finances_ Rework.xlsx ({date_cell})",
        })

    return records


def extract_tracker_income_transactions(
    ws_kd, ws_kak, keaton_accounts: dict, katherine_accounts: dict,
    income_cat_id: int | None, cutoff_date: date
) -> list:
    """
    Extract monthly income transactions from KD/KAK Ongoing Trackers
    for months before cutoff_date (i.e. before paystubs start Feb 2025).

    Sheet structure:
      Row 6  = month headers (datetime objects: Sep 2024, Oct 2024, ...)
      Row 20 = monthly Income totals for each column

    Returns list of (username, transaction_dict) tuples.
    """
    results = []

    for username, ws, account_ids in [
        ("keaton",    ws_kd,  keaton_accounts),
        ("katherine", ws_kak, katherine_accounts),
    ]:
        checking_id = account_ids.get("checking")
        if not checking_id:
            print(f"  [SKIP] {username}: no 'checking' account found in app -- create it first")
            continue
        if not income_cat_id:
            print(f"  [SKIP] No income category found in app -- monthly transactions skipped")
            break

        month_row  = [ws.cell(6,  col).value for col in range(1, 40)]
        income_row = [ws.cell(20, col).value for col in range(1, 40)]

        for col_idx in range(1, len(month_row)):
            month_dt = month_row[col_idx]
            if not isinstance(month_dt, datetime):
                continue

            month_date = month_dt.date()
            snap_date  = last_day_of_month(month_date.year, month_date.month)
            if snap_date >= cutoff_date:
                continue

            amount = _num(income_row[col_idx])
            if amount is None or amount <= 0:
                continue

            results.append((username, {
                "date":        snap_date.isoformat(),
                "amount":      amount,
                "description": f"Monthly income {month_date.strftime('%b %Y')} (xlsx import)",
                "category_id": income_cat_id,
                "account_id":  checking_id,
                "is_verified": True,
            }))

    return results


def extract_student_loan_payments(
    ws_kd, ws_kak,
    keaton_accounts: dict, katherine_accounts: dict,
    loan_cat_id: int | None, cutoff_date: date
) -> list:
    """
    Extract monthly student loan payment transactions from KD/KAK Ongoing Trackers.

    Only imports months BEFORE cutoff_date (the Google Sheets sync start date).
    After cutoff, the monthly spending Google Sheet is the source of truth.

    Sheet rows (1-indexed):
      KD  row 132 = Keaton's combined student loan payment line
      KAK row 108 = Katherine's combined student loan payment line
      Row 6       = month headers (datetime objects)

    Payments are posted as expense transactions against the user's student_loan account.
    The student_loan account MUST exist in the app first. The user should manually
    add a balance snapshot for the starting balance (Sep 2024).
    """
    results = []
    loan_row_nums = {"keaton": 132, "katherine": 108}

    for username, ws, account_ids in [
        ("keaton",    ws_kd,  keaton_accounts),
        ("katherine", ws_kak, katherine_accounts),
    ]:
        loan_acct_id = account_ids.get("student_loan")
        if not loan_acct_id:
            print(f"  [SKIP] {username}: no 'student_loan' account found in app -- create it first")
            continue
        if not loan_cat_id:
            print(f"  [SKIP] No 'Student Loans' expense category found -- loan transactions skipped")
            break

        row_num = loan_row_nums[username]
        month_row = [ws.cell(6,       col).value for col in range(1, 40)]
        loan_row  = [ws.cell(row_num, col).value for col in range(1, 40)]

        for col_idx in range(1, len(month_row)):
            month_dt = month_row[col_idx]
            if not isinstance(month_dt, datetime):
                continue

            month_date = month_dt.date()
            snap_date  = last_day_of_month(month_date.year, month_date.month)
            if snap_date >= cutoff_date:
                continue  # covered by Google Sheets sync

            amount = _num(loan_row[col_idx])
            if amount is None or amount <= 0:
                continue

            results.append((username, {
                "date":        snap_date.isoformat(),
                "amount":      abs(amount),
                "description": f"Student loan payment {month_date.strftime('%b %Y')} (xlsx import)",
                "category_id": loan_cat_id,
                "account_id":  loan_acct_id,
                "is_verified": True,
            }))

    return results


def extract_bonus_transactions(
    ws_kd, ws_kak,
    keaton_accounts: dict, katherine_accounts: dict,
    bonus_cat_id: int | None,
) -> list:
    """
    Extract bonus income transactions from KD/KAK Ongoing Trackers.

    Bonuses are manual input cells in the Rework sheet and are NOT captured in the
    monthly spending Google Sheets, so we import ALL months with a non-zero value
    (no date cutoff applied).

    Sheet rows (1-indexed):
      KD  row 30       = Keaton bonus income
      KAK rows 26, 27  = Katherine bonus income (two lines, e.g. one per employer)
      Row 6            = month headers (datetime objects)

    Bonus transactions are posted as income against the user's checking account.
    """
    results = []
    bonus_row_nums = {"keaton": [30], "katherine": [26, 27]}

    for username, ws, account_ids in [
        ("keaton",    ws_kd,  keaton_accounts),
        ("katherine", ws_kak, katherine_accounts),
    ]:
        checking_id = account_ids.get("checking")
        if not checking_id:
            print(f"  [SKIP] {username}: no 'checking' account found -- bonus transactions skipped")
            continue
        if not bonus_cat_id:
            print(f"  [SKIP] No 'Bonus' income category found -- bonus transactions skipped")
            break

        month_row = [ws.cell(6, col).value for col in range(1, 40)]

        for row_num in bonus_row_nums[username]:
            # Read the row label from column A so the description is descriptive
            row_label = ws.cell(row_num, 1).value
            label = str(row_label).strip() if row_label else "Bonus"
            # Normalize label to something clean
            if not label or label.lower() in ("none", ""):
                label = "Bonus"

            bonus_row = [ws.cell(row_num, col).value for col in range(1, 40)]

            for col_idx in range(1, len(month_row)):
                month_dt = month_row[col_idx]
                if not isinstance(month_dt, datetime):
                    continue

                amount = _num(bonus_row[col_idx])
                if amount is None or amount <= 0:
                    continue

                month_date = month_dt.date()
                snap_date  = last_day_of_month(month_date.year, month_date.month)

                results.append((username, {
                    "date":        snap_date.isoformat(),
                    "amount":      amount,
                    "description": f"{label} {month_date.strftime('%b %Y')} (xlsx import)",
                    "category_id": bonus_cat_id,
                    "account_id":  checking_id,
                    "is_verified": True,
                }))

    return results


# -----------------------------------------------------------------------------
# API post helpers
# -----------------------------------------------------------------------------

def post_paystub(api_base: str, token: str, payload: dict, dry_run: bool) -> bool:
    if dry_run:
        print(f"    [DRY RUN] POST /paystubs  pay_date={payload['pay_date']}"
              f"  gross={payload.get('gross_pay')}  net={payload.get('net_pay')}")
        return True
    resp = requests.post(
        f"{api_base}/paystubs",
        json=payload,
        headers={"Authorization": f"Bearer {token}"},
        timeout=10,
    )
    if resp.status_code in (200, 201):
        print(f"    OK Created paystub {payload['pay_date']}")
        return True
    print(f"    FAIL paystub {payload['pay_date']}: {resp.status_code} {resp.text[:120]}")
    return False


def post_snapshot(api_base: str, token: str, payload: dict, dry_run: bool) -> bool:
    if dry_run:
        print(f"    [DRY RUN] POST /balance-snapshots  acct_id={payload['account_id']}"
              f"  date={payload['date']}  balance={payload['balance']}")
        return True
    resp = requests.post(
        f"{api_base}/balance-snapshots",
        json=payload,
        headers={"Authorization": f"Bearer {token}"},
        timeout=10,
    )
    if resp.status_code in (200, 201):
        print(f"    OK Snapshot acct={payload['account_id']} {payload['date']} ${payload['balance']:,.2f}")
        return True
    print(f"    FAIL snapshot acct={payload['account_id']} {payload['date']}: {resp.status_code} {resp.text[:120]}")
    return False


def post_transaction(api_base: str, token: str, payload: dict, dry_run: bool) -> bool:
    if dry_run:
        print(f"    [DRY RUN] POST /transactions  date={payload['date']}"
              f"  ${payload['amount']:,.2f}  {payload['description'][:50]}")
        return True
    resp = requests.post(
        f"{api_base}/transactions",
        json=payload,
        headers={"Authorization": f"Bearer {token}"},
        timeout=10,
    )
    if resp.status_code in (200, 201):
        print(f"    OK Transaction {payload['date']} ${payload['amount']:,.2f}")
        return True
    print(f"    FAIL transaction {payload['date']}: {resp.status_code} {resp.text[:120]}")
    return False


# -----------------------------------------------------------------------------
# Utilities
# -----------------------------------------------------------------------------

def _num(v):
    """Return float if numeric, else None."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        try:
            return float(v.replace(",", "").strip())
        except ValueError:
            return None
    return None


def load_config() -> dict:
    """Load config.json if it exists, otherwise return defaults."""
    if CONFIG_FILE.exists():
        with open(CONFIG_FILE) as f:
            cfg = json.load(f)
        # Merge with defaults so new keys are always present
        return {**DEFAULT_CONFIG, **cfg}
    return dict(DEFAULT_CONFIG)


# -----------------------------------------------------------------------------
# Main
# -----------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Import historical data from KK Finances Excel")
    parser.add_argument("--dry-run",           action="store_true", help="Preview without writing")
    parser.add_argument("--user",              choices=["keaton", "katherine"], help="Import one user only")
    parser.add_argument("--skip-paystubs",     action="store_true")
    parser.add_argument("--skip-snapshots",    action="store_true")
    parser.add_argument("--skip-transactions", action="store_true", help="Skip monthly income transactions")
    parser.add_argument("--skip-loans",        action="store_true", help="Skip student loan payment transactions")
    parser.add_argument("--skip-bonuses",      action="store_true", help="Skip bonus income transactions")
    args = parser.parse_args()

    cfg      = load_config()
    api_base = cfg["api_base"]
    xlsx_path = cfg["xlsx_path"]

    print(f"\n{'='*60}")
    print(f"  FinanceTrack Historical Import")
    print(f"  Excel:  {xlsx_path}")
    print(f"  API:    {api_base}")
    print(f"  Mode:   {'DRY RUN' if args.dry_run else 'LIVE'}")
    if args.user:
        print(f"  User:   {args.user} only")
    print(f"{'='*60}\n")

    # -- Load workbook ---------------------------------------------------------
    if not Path(xlsx_path).exists():
        print(f"ERROR: Excel file not found: {xlsx_path}")
        print(f"  Expected at: {xlsx_path}")
        sys.exit(1)

    print("Loading workbook...")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws_keaton = wb["Keaton PayStub"]
    ws_kath   = wb["Katherine PayStub"]
    ws_kd     = wb["KD Ongoing Tracker"]
    ws_kak    = wb["KAK Ongoing Tracker"]
    print("  OK Loaded 4 sheets\n")

    # -- KEATON ----------------------------------------------------------------
    token_k = None
    keaton_account_ids = {}

    if args.user in (None, "keaton"):
        print("--- KEATON ----------------------------------------------")

        if not args.dry_run:
            try:
                token_k = login(api_base, cfg["keaton_username"], cfg["keaton_password"])
                print("  OK Logged in as keaton")
            except Exception as e:
                print(f"  FAIL Login failed: {e}")
                sys.exit(1)

            # Auto-discover accounts and income category
            keaton_account_ids = discover_accounts(api_base, token_k)
            print(f"  OK Found {len(keaton_account_ids)} accounts: {list(keaton_account_ids.keys())}")
        else:
            # Dry run -- show placeholder IDs so output is readable
            keaton_account_ids = {"checking": "?", "savings": "?", "hysa": "?", "401k": "?", "ira": "?"}

        # Paystubs
        if not args.skip_paystubs:
            print("\n  [Paystubs]")
            stubs = extract_keaton_paystubs(ws_keaton)
            print(f"  Found {len(stubs)} paystub records")
            existing = set() if args.dry_run else get_existing_paystub_dates(api_base, token_k)
            ok = skip = 0
            for stub in stubs:
                if stub["pay_date"] in existing:
                    print(f"    [SKIP] {stub['pay_date']} already exists")
                    skip += 1
                else:
                    post_paystub(api_base, token_k, stub, args.dry_run) and (ok := ok + 1)
            print(f"  -> {ok} posted, {skip} skipped")

        # Balance snapshots
        if not args.skip_snapshots:
            print("\n  [Balance Snapshots]")
            snaps = extract_keaton_snapshots(ws_keaton, keaton_account_ids)
            if not snaps:
                print("  No snapshots found (accounts may not exist in app yet -- create them first)")
            else:
                print(f"  Found {len(snaps)} snapshot records")
                ok = skip = 0
                for snap in snaps:
                    acct_id = snap["account_id"]
                    existing = set() if args.dry_run else get_existing_snapshot_dates(api_base, token_k, acct_id)
                    if snap["date"] in existing:
                        print(f"    [SKIP] acct={acct_id} {snap['date']} already exists")
                        skip += 1
                    else:
                        post_snapshot(api_base, token_k, snap, args.dry_run) and (ok := ok + 1)
                print(f"  -> {ok} posted, {skip} skipped")

    # -- KATHERINE -------------------------------------------------------------
    token_kat = None
    katherine_account_ids = {}

    if args.user in (None, "katherine"):
        print("\n--- KATHERINE -------------------------------------------")

        if not args.dry_run:
            try:
                token_kat = login(api_base, cfg["katherine_username"], cfg["katherine_password"])
                print("  OK Logged in as katherine")
            except Exception as e:
                print(f"  FAIL Login failed: {e}")
                sys.exit(1)

            katherine_account_ids = discover_accounts(api_base, token_kat)
            print(f"  OK Found {len(katherine_account_ids)} accounts: {list(katherine_account_ids.keys())}")
        else:
            katherine_account_ids = {"checking": "?", "savings": "?", "401k": "?", "ira": "?"}

        if not args.skip_paystubs:
            print("\n  [Paystubs]")
            stubs = extract_katherine_paystubs(ws_kath)
            print(f"  Found {len(stubs)} paystub records")
            existing = set() if args.dry_run else get_existing_paystub_dates(api_base, token_kat)
            ok = skip = 0
            for stub in stubs:
                if stub["pay_date"] in existing:
                    print(f"    [SKIP] {stub['pay_date']} already exists")
                    skip += 1
                else:
                    post_paystub(api_base, token_kat, stub, args.dry_run) and (ok := ok + 1)
            print(f"  -> {ok} posted, {skip} skipped")

        if not args.skip_snapshots:
            print("\n  [Balance Snapshots]")
            print("  Katherine's balance sheet has no populated balance data in the spreadsheet -- skipping.")

    # -- MONTHLY INCOME TRANSACTIONS (Sep 2024 - Jan 2025) --------------------
    if not args.skip_transactions:
        print("\n--- MONTHLY INCOME TRANSACTIONS (from tracker) ----------")

        # Get income category -- use keaton's token (categories are global)
        income_cat_id = None
        if not args.dry_run:
            active_token = token_k or token_kat
            if active_token:
                income_cat_id = discover_income_category(api_base, active_token)
                if income_cat_id:
                    print(f"  OK Income category id={income_cat_id}")
                else:
                    print("  WARN No income category found (kind=income) -- monthly transactions skipped")

        # Paystubs start Feb 2025; import only months before that
        cutoff = date(2025, 2, 1)
        print(f"  Importing months before {cutoff} (Sep 2024 - Jan 2025)")

        tx_list = extract_tracker_income_transactions(
            ws_kd, ws_kak,
            keaton_account_ids, katherine_account_ids,
            income_cat_id, cutoff
        )
        print(f"  Found {len(tx_list)} transaction records")

        for username, tx in tx_list:
            if args.user and args.user != username:
                continue
            token = token_k if username == "keaton" else token_kat
            if args.dry_run:
                print(f"    [{username}] [DRY RUN] {tx['date']}  ${tx['amount']:,.2f}  {tx['description'][:50]}")
            else:
                post_transaction(api_base, token, tx, False)

    # -- STUDENT LOAN PAYMENTS (Sep 2024 - Jan 2025) ---------------------------
    if not args.skip_loans:
        print("\n--- STUDENT LOAN PAYMENTS (from tracker) ----------------")
        print("  NOTE: 'Student Loans' account must exist in app for each user.")
        print("        Payments are only imported for Sep 2024 - Jan 2025")
        print("        (Google Sheets covers Feb 2025 onward).")

        loan_cat_id = None
        if not args.dry_run:
            active_token = token_k or token_kat
            if active_token:
                loan_cat_id = discover_category_by_name(api_base, active_token, "Student Loans")
                if loan_cat_id:
                    print(f"  OK 'Student Loans' category id={loan_cat_id}")
                else:
                    print("  WARN 'Student Loans' category not found -- loan payments skipped")
        else:
            loan_cat_id = "?"

        cutoff = date(2025, 2, 1)
        loan_txns = extract_student_loan_payments(
            ws_kd, ws_kak,
            keaton_account_ids, katherine_account_ids,
            loan_cat_id, cutoff
        )
        print(f"  Found {len(loan_txns)} loan payment records")

        for username, tx in loan_txns:
            if args.user and args.user != username:
                continue
            token = token_k if username == "keaton" else token_kat
            if args.dry_run:
                print(f"    [{username}] [DRY RUN] {tx['date']}  ${tx['amount']:,.2f}  {tx['description'][:60]}")
            else:
                existing = get_existing_transaction_keys(api_base, token)
                key = (tx["date"], tx["description"])
                if key in existing:
                    print(f"    [SKIP] {tx['date']} {tx['description'][:50]} already exists")
                else:
                    post_transaction(api_base, token, tx, False)

    # -- BONUS INCOME (all months with a value) --------------------------------
    if not args.skip_bonuses:
        print("\n--- BONUS INCOME (from tracker) -------------------------")
        print("  Bonuses are manual entries in Rework, not in monthly spending sheets.")
        print("  Importing all months with a non-zero bonus value.")

        bonus_cat_id = None
        if not args.dry_run:
            active_token = token_k or token_kat
            if active_token:
                bonus_cat_id = discover_category_by_name(api_base, active_token, "Bonus")
                if bonus_cat_id:
                    print(f"  OK 'Bonus' category id={bonus_cat_id}")
                else:
                    print("  WARN 'Bonus' income category not found -- bonus transactions skipped")
        else:
            bonus_cat_id = "?"

        bonus_txns = extract_bonus_transactions(
            ws_kd, ws_kak,
            keaton_account_ids, katherine_account_ids,
            bonus_cat_id,
        )
        print(f"  Found {len(bonus_txns)} bonus records")

        for username, tx in bonus_txns:
            if args.user and args.user != username:
                continue
            token = token_k if username == "keaton" else token_kat
            if args.dry_run:
                print(f"    [{username}] [DRY RUN] {tx['date']}  ${tx['amount']:,.2f}  {tx['description'][:60]}")
            else:
                existing = get_existing_transaction_keys(api_base, token)
                key = (tx["date"], tx["description"])
                if key in existing:
                    print(f"    [SKIP] {tx['date']} {tx['description'][:50]} already exists")
                else:
                    post_transaction(api_base, token, tx, False)

    print(f"\n{'='*60}")
    print("  Import complete.")
    if args.dry_run:
        print("  This was a DRY RUN -- nothing was written to the database.")
        print("  Remove --dry-run to perform the actual import.")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
